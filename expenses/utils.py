# expenses/utils.py
import io
import os
import zipfile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.core.files.storage import default_storage

# Headers and widths in the exact order you requested
EXPORT_HEADERS = [
    "id",                 # 1
    "job code/number",    # 2 -> project_code or project name
    "description",        # 3
    "quantity",           # 4
    "price",              # 5
    "total",              # 6
    "account",            # 7
    "subaccount",         # 8
    "user",               # 9
    "receipt",            # 10 (relative path inside the zip)
]
COL_WIDTHS = [6, 20, 40, 12, 12, 12, 16, 18, 18, 28]


def build_export(expenses_qs):
    """
    Returns an in-memory ZIP containing:
      - expenses.xlsx with the requested columns
      - receipts/ folder with the images referenced in the sheet
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "expenses"

    # Headers + widths
    ws.append(EXPORT_HEADERS)
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    mem = io.BytesIO()
    zf = zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED)

    def add_receipt_to_zip(storage_path, zip_rel_path):
        with default_storage.open(storage_path, "rb") as fh:
            zf.writestr(zip_rel_path, fh.read())

    # Build rows and copy first receipt (if any)
    for e in expenses_qs.select_related('project', 'created_by').prefetch_related('receipts').order_by('date', 'id'):
        job_code = e.project_code or (e.project.name if getattr(e, "project", None) else "")

        first_receipt_rel = ""
        receipts = list(e.receipts.all())
        if receipts:
            r0 = receipts[0]
            storage_path = r0.image.name
            filename = os.path.basename(storage_path)
            zip_path = f"receipts/{e.id}_{filename}"
            add_receipt_to_zip(storage_path, zip_path)
            first_receipt_rel = zip_path

        ws.append([
            e.id,
            job_code,
            e.description or "",
            float(e.quantity or 0),
            float(e.unit_price or 0),
            float(e.total or 0),
            e.account or "",
            e.subaccount or "",
            (e.created_by.username if e.created_by else ""),
            first_receipt_rel,
        ])

    # Save workbook inside the zip
    xls_bytes = io.BytesIO()
    wb.save(xls_bytes)
    xls_bytes.seek(0)
    zf.writestr("expenses.xlsx", xls_bytes.read())

    zf.close()
    mem.seek(0)
    return mem
